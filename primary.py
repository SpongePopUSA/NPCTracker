# primary.py
# This program keeps track of NPC's in the Elcria Campaign

from decorator import *
from id import IDTracker
from npc import NPC
from event import Event, Date, Month
from openpyxl import *
from openpyxl.worksheet.worksheet import Worksheet
import time

# Initialize workbook and worksheet
wb:Workbook = load_workbook("PartyStatus.xlsx")
ws:Worksheet = wb.active
# Bounds of npc table
npcBounds = {
    "topRow"    : 4,
    "leftCol"   : 2,
    "rightCol"  : 3
}
# Bounds of event table
eventBounds = {
    "topRow"    : 4,
    "leftCol"   : 5,
    "rightCol"  : 11
}
# List of events
eventsList:list[Event] = []
# List of NPC's
npcList:list[NPC] = []

# Populate NPC and event lists
def initializeData ():
    print("Initializing NPC's...\n")
    # Iterate across NPC table
    for val in ws.iter_rows(min_row = npcBounds["topRow"], min_col = npcBounds["leftCol"], max_col = npcBounds["rightCol"], values_only = True):
        # If the row has data
        if val[0] != None:
            # Store name
            name = str(val[0])
            # Create new NPC
            new_NPC = NPC(name = name)
            # Add NPC to list
            npcList.append(new_NPC)
            # Display progress with time delay for aesthetics
            print(f"Initialized: {new_NPC}")
            time.sleep(.03)
        # If the row is empty, skip iteration
        else: continue
    print("\nDone.\n")

    print("Initializing events...\n")
    # Iterate across event table
    for val in ws.iter_rows(min_row = eventBounds["topRow"], min_col = eventBounds["leftCol"], max_col = eventBounds["rightCol"], values_only = True):
        # If the row has data
        if val[0] != None:
            # Store title
            title = val[0]
            # Store day
            d = int(val[1])
            # Store month
            m = int(val[2])
            # Store year
            y = int(val[3])
            # Store list of affected/delta pairs
            ad_pairs:list[str] = str(val[4]).split(",")
            # List of (affected, delta) tuples
            ad_tuples:list[tuple[str, int]] = []
            # Populate ad_tuples
            for pair in ad_pairs:
                # If NPC's were affected by this event
                if pair != "None":
                    # String list of size 2: [ID of affected NPC, Change in reputation score]
                    split_pair = pair.split("/")
                    # String representation of character's ID
                    id_string = split_pair[0] + 'n'
                    # Integer representing the change in character's reputation score
                    delta = int(split_pair[1])
                    # Add new tuple to list
                    ad_tuples.append((id_string, delta))

            # List of id_strings of killed NPC's
            killed_list:list[str] = []
            # Populate killed_list
            for npc_id in str(val[5]).split(','):
                # If NPC's were killed in this event
                if npc_id != "None":
                    # Add string representation of NPC ID to list
                    killed_list.append(npc_id.strip() + 'n')

            # Create a new event with info from this row
            new_event = Event(title = title, date = Date(d, m, y), involved = ad_tuples, killed = killed_list)
            # Add event to list
            eventsList.append(new_event)
            # Display progress 
            print(f"Initialized: {new_event}")
            # Time delay for aesthetics
            time.sleep(.03)
        # If the row is empty, skip iteration
        else: continue
    print("\nDone.")
# Display naviagation options
def displayMainMenu ():
    # Dictionary representing options in the menu and the functions that they call
    options = {
    # key : (function, description)
        1 : (showNPCInfo, "Get NPC info"),
        2 : (showEventInfo, "Get event info"),
        3 : (addNewNPC, "Add NPC"),
        4 : (addNewEvent, "Add event")
    }
    # Initialize main menu with header
    menu_message = "Welcome to Elcria's NPC Tracker!\n\nPlease select from the following:\n\n"
    # Fill in the menu with descriptions from the options dictionary
    for key in options:
        menu_message = menu_message + f"     {key}  :  {options[key][1]}\n"
    # Add exit option
    menu_message = menu_message + f"\n    -1  :  Exit\n"
    # Variable for user's input
    choice = None
    # While the user has not made a valid choice
    while choice == None:
        try:
            # Display main menu
            print(menu_message)
            # Take input
            choice:int = int(input("Selection: "))
            # If the user chooses to exit
            if choice == -1: return
            # Decorate the chosen function with a line of hashtags
            decorate(func = options[choice][0], item = '#')
            # Reset choice
            choice = None
            # Next iteration
            continue
        # If the input is not a valid type
        except ValueError:
            print("Please input an integer.")
            # Reset choice
            choice = None
            # Next iteration
            continue
        # If the input is not a valid option
        except KeyError:
            print("Please input a valid option.")
            # Reset choice
            choice = None
            # Next iteration
            continue
        # If some other error occurred
        except Exception as error:
            print(f"Something went wrong, please try again.\nError: {error}")
            # Reset choice
            choice = None
            # Next iteration
            continue
    # Gap
    print("\n")
# Search for an NPC and display its info to the user           
def showNPCInfo ():
    # String representing all NPC's and their ID numbers
    help_table = ""
    # Populate help_table
    for npc in npcList:
        # Store name
        name = npc.name
        # Store ID number with leading zeros for single-digit ID's
        num = str(npc.ident.num).zfill(2)
        # Add line to help_table
        help_table = help_table + f"   {num}  :  {name}\n"
    # Intro message
    print("Welcome to the NPC search!\nType 'EXIT' to quit or 'HELP' for reference.\n")
    # Variable for user's input
    choice = None
    # While the user has not made a valid choice
    while choice == None:
        try:
            # Prompt user for input
            choice = input("Please input the NPC's ID without holder tag: ").strip().upper()
            # Gap
            print(" ")
            # If the user chooses to exit
            if choice == "EXIT": return
            # If user asks for reference table
            elif choice == "HELP":
                print(help_table)
                # Reset choice
                choice = None
                # Next iteration
                continue
            else:
                # The character associated with the ID
                # Cast from string to int back to string to remove leading zeros
                character:NPC = IDTracker.findByID(str(int(choice)) + 'n')
                # Display character summary
                print(character.getSummary())
                # Reset choice
                choice = None
                # Next iteration
                continue
        # If the input is not a valid ID
        except AttributeError:
            print("Please input a valid ID.")
            # Reset choice
            choice = None
            # Next iteration
            continue
        # If some other error occurred
        except Exception as error:
            print(f"Something went wrong, please try again.")
            # Reset choice
            choice = None
            # Next iteration
            continue
# Search for an event and display its info to the user
# Syntax is largely the same as showNPCInfo() (Possible refactor?)           
def showEventInfo ():
    # String representing all eventss and their ID numbers
    help_table = ""
    # Populate help_table
    for event in eventsList:
        # Store title
        title = event.title
        # Store ID number with leading zeros for single-digit ID's
        num = str(event.ident.num).zfill(2)
        # Add line to help_table
        help_table = help_table + f"   {num}  :  {title}\n"
    # Intro message
    print("Welcome to the NPC search!\nType 'EXIT' to quit or 'HELP' for reference.\n")
    # Variable for user's input
    choice = None
    # While the user has not made a valid choice
    while choice == None:
        try:
            # Prompt user for input
            choice = input("Please input the event's ID without holder tag: ").strip().upper()
            # If the user chooses to exit
            if choice == "EXIT": return
            # If user asks for reference table
            elif choice == "HELP":
                print(help_table)
                # Reset choice
                choice = None
                # Next iteration
                continue
            else:
                # The character associated with the ID
                # Cast from string to int back to string to remove leading zeros
                event:Event = IDTracker.findByID(str(int(choice)) + 'e')
                # Display event summary
                print( event.getSummary() )
                # Reset choice
                choice = None
                # Next iteration
                continue
        # If the input is not a valid ID
        except AttributeError:
            print("Please input a valid ID.")
            # Reset choice
            choice = None
            # Next iteration
            continue
        # If some other error occurred
        except Exception as error:
            print("Something went wrong, please try again.")
            # Reset choice
            choice = None
            # Next iteration
            continue
# Create a new NPC
def addNewNPC ():
    # Intro message
    print("Welcome to the NPC creation wizard!\nType 'EXIT' at any time to return to the main menu.\n")
    # Prompt for NPC name
    character_name = None
    # While user has not submitted a valid input
    while character_name == None:
        # Prompt for the name of the NPC
        character_name = input("Please input the NPC's name: ").strip()
        # If user wrote nothing
        if character_name == "":
            print("Please input a name.")
            # Reset choice
            character_name = None
            # Next iteration
            continue
        # If the user chooses to exit
        elif character_name == "EXIT": return
    # Create a new NPC
    new_npc = NPC(name = character_name)
    # Add new NPC to list
    npcList.append(new_npc)

    print(f"\nAdded new NPC: {character_name}")
# Create a new event
def addNewEvent ():
    # Prompt the user to input the event date
    def promptDate () -> Date:
        # Day of the month
        event_day = None
        while event_day == None:
            event_day = input("Please input the day of the month: ")
            if event_day.lstrip('-').isnumeric() and 1 <= int(event_day) <= 27:
                event_day = int(event_day)
            elif event_day== "EXIT":
                return
            else:
                print("Please input a valid day.")
                event_day = None
        event_month = None
        while event_month == None:
            event_month = input("Please input the month: ")
            if event_month.isnumeric():
                if 1 <= int(event_month) <= 16:
                    event_month = Month(int(event_month))
                else:
                    print("Please input a valid month.")
                    event_month = None
            elif event_month.strip().upper() in Month.__members__:
                event_month = Month[event_month.strip().upper()]
            elif event_title == "EXIT":
                return
            else:
                print("Please input a valid month.")
                event_month = None
        event_year = None
        while event_year == None:
            event_year = input("Please input the year: ")
            if event_year.lstrip('-').isnumeric() and 1 <= int(event_day) <= 27:
                event_year = int(event_year)
            elif event_year == "EXIT":
                return
            else:
                print("Please input a valid year.")
                event_year = None
        
        result = Date(event_day, event_month, event_year)
        return result
    # Prompt the user to generate a list of NPC/delta pairs
    def promptInvolvedList () -> list[tuple[str, int]]:
        # List of ID/delta pairs
        involved_list:list[tuple[str, int]] = []
        # Boolean: True = Continue prompting, False = Done prompting
        adding_pair:bool = True

        while adding_pair:
            # String representation of the character's ID 
            character_id_string = ""
            # Change to character's reputation score
            delta:int = None
            # String representation of already added characters, will be printed after all characters are added
            involved_string = "Involved: "

            # Iterate across list of added NPC's
            for inv in involved_list:
                # Actual character object
                character:NPC = IDTracker.findByID(inv[0])
                # Change to reputation score, represented by a string to append +/- sign
                delta_string:str = ""
                # If the change to reputation score is positive or zero
                if inv[1] >= 0:
                    # Add a plus sign to the string
                    delta_string = '+' + str(inv[1])
                # If the change to reputation score is negative
                else:
                    # Cast the int to a string (- sign will already be included)
                    delta_string = str(inv[1])
                # Add this character to the string representation
                involved_string = involved_string + character.name + ' (' + delta_string + ')' + ", "
            # Remove tailing comma
            involved_string = involved_string.rstrip(", ")
            # If no characters have been added
            if len(involved_list) == 0:
                print("\nNo involved NPC's.\n")
            # If 1+ characters have already been added
            else:
                # Show list of already added characters
                print(f"\n{involved_string}\n")
            # If the user wants to add more characters
            choice = 'Y'
            # While the user is still adding characters
            while choice == 'Y':
                # Ask the user if they would like to add another NPC
                choice = input("Add involved NPC? (Y/N) ").strip().upper()
                match choice:
                    case 'Y':
                        # Name of character to be added
                        character_name = None
                        # While the user has not submitted a name
                        while character_name == None:
                            # Prompt for name
                            character_name = input("Please input the name of the involved NPC: ").strip()
                            # Character object of chosen NPC
                            character:NPC = None
                            # If user chooses to exit
                            if character_name == "EXIT": return
                            # Set the name to lower case
                            character_name = character_name.lower()
                            # Check that user has not already submitted the same character
                            if involved_string.lower().find(character_name) != -1:
                                print("Duplicates are not allowed.")
                                character_name = None
                                continue
                            for n in npcList:
                                if n.name.lower() == character_name:
                                    character = n
                                else:
                                    continue
                            if character != None:
                                character_id_string = str(character.ident)
                            else:
                                print("Please input a valid NPC name.")
                                character_name = None
                                continue
                        while delta == None:
                            delta_input = input("Please input the change in reputation score for this NPC: ").strip()
                            if delta_input.lstrip('-').isnumeric():
                                delta = int(delta_input)
                            else:
                                print("Please input a valid integer.")
                                delta = None
                        break
                    case 'N':
                        adding_pair = False
                        break
                    case "EXIT":
                        return -1
                    case _:
                        print("Please input a valid option")
                        choice = 'Y'
            
            if character_id_string != '' and delta != None:
                involved_list.append((character_id_string, delta))

        return involved_list
    # Prompt the user to generate a list of killed NPC's
    def promptKilledList () -> list[str]:
        killed_list:list[str] = []

        adding_npc:bool = True
        while adding_npc:
            character_id_string = ""

            killed_string = ""
            for k in killed_list:
                character:NPC = IDTracker.findByID(k)
                killed_string = killed_string + character.name + ", "
            killed_string = killed_string.rstrip(", ")
            if len(killed_list) == 0:
                print("\nNo killed NPC's.\n")
            else:
                print("\nKilled: " + killed_string + "\n")
            
            choice = 'Y'
            while choice == 'Y':
                choice = input("Add killed NPC? (Y/N) ").strip().upper()
                match choice:
                    case 'Y':
                        character_name = None
                        while character_name == None:
                            character_name = input("Please input the name of the killed NPC: ").strip().lower()
                            # Check that user has not already submitted the same character
                            if killed_string.lower().find(character_name) != -1:
                                print("Duplicates are not allowed.")
                                character_name = None
                                continue
                            character:NPC = None
                            for n in npcList:
                                if n.name.lower() == character_name:
                                    character = n
                                else:
                                    continue
                            if character != None:
                                character_id_string = str(character.ident)
                            else:
                                print("Please input a valid NPC name.")
                                character_name = None
                        break
                    case 'N':
                        adding_npc = False
                        break
                    case "EXIT":
                        return -1
                    case _:
                        print("Please input a valid option")
                        choice = 'Y'
            
            if (character_id_string != ''):
                killed_list.append(character_id_string)

        return killed_list

    # Intro message
    print("Welcome to the event creation wizard!\nType 'EXIT' at any time to return to the main menu.\n")
    # Variable for title input
    event_title = None
    # While user title input is invalid
    while event_title == None:
        # Take input
        event_title = input("Please input a single-sentence description of the event: ").strip()
        # If user chooses to exit
        if event_title == "EXIT": return
        # If description is only whitespace
        elif event_title == "":
            print("Please input a description.")
            # Reset input
            event_title = None
            # Nest iteration
            continue
    # Prompt for the date of the event
    event_date = promptDate()
    # Prompt for list of NPC/delta pairs
    delta_pairs = promptInvolvedList()
    # Prompt for list of killed NPC's
    killed_npcs = promptKilledList()

    # Create new event
    new_event = Event(event_title, event_date, delta_pairs, killed_npcs)
    # Add event to list
    eventsList.append(new_event)

    print(f"\nAdded event: {str(new_event)}.")
# Update workbook with current NPC's and Events
def update(update_npcs:bool = True, update_events:bool = True):
    # Save backup
    wb.save("PartyStatus_Backup.xlsx")

    print("\nUpdating NPC's...\n")
    # Fill NPC data
    for character in npcList:
        # Get appropriate row from id
        r = character.ident.num + 3
        # Update npc cells
        ws.cell(r, npcBounds["leftCol"]).value = character.name
        ws.cell(r, npcBounds["rightCol"]).value = character.ident.num
    print("\nDone.\n")

    print("Updating Events...\n")
    # Update event cells
    for event in eventsList:
        # Get appropriate row from id
        r = event.ident.num + 3
        ws.cell(r, eventBounds["leftCol"]).value = event.title
        ws.cell(r, eventBounds["leftCol"] + 1).value = event.date.day_num
        ws.cell(r, eventBounds["leftCol"] + 2).value = event.date.month
        ws.cell(r, eventBounds["leftCol"] + 3).value = event.date.year
        # Update affected/delta cell
        # If no NPC's were involved
        if len(event.involved) == 0:
            ws.cell(r, eventBounds["leftCol"] + 4).value = "None"
        else:
            # Variable to store string representation of pair list
            ad_string = ""
            # For each NPC involved
            for ad_tuple in event.involved:
                # Find the NPC
                character:NPC = IDTracker.findByID(ad_tuple[0])
                # Store ID number
                id_num = str(character.ident.num)
                # Store the change in reputation
                delta = str(ad_tuple[1])
                # Add the info to the string
                ad_string = ad_string + id_num + "/" + delta + ","
            # Remove comma at the end of ad_string
            ad_string = ad_string[:-1]
            # Update the cell with list of pairs
            ws.cell(r, eventBounds["leftCol"] + 4).value = ad_string
        # Update killed cell
        # If no NPC's were killed
        if len(event.killed) == 0:
            ws.cell(r, eventBounds["leftCol"] + 5).value = "None"
        else:
            # Variable to store string representation of killed list
            killed_string = ""
            # For each NPC killed
            for id_string in event.killed:
                # Find the NPC
                character:NPC = IDTracker.findByID(id_string)
                # Store ID number
                id_num = str(character.ident.num)
                # Add the ID number to the string
                killed_string = killed_string + id_num + ","
            # Remove comma at the end of killed_string
            killed_string = killed_string[:-1]
            # Update the cell with the ID number list
            ws.cell(r, eventBounds["leftCol"] + 5).value = killed_string
        # Update the id cell
        ws.cell(r, eventBounds["leftCol"] + 6).value = event.ident.num
    print("\nDone.\n")
# Save changes to workbook and close
def saveAndClose ():
    print("Saving...")
    wb.save("PartyStatus.xlsx")
    wb.close()
    print("Done.")

def main():
    initializeData()
    displayMainMenu()
    update()
    saveAndClose()

decorate(func = main)